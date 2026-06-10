"""Web dashboard for the copytrader bots.

Reads the state_<name>.json snapshots written each cycle by bot.py and renders one
tracker at a time (switchable via a toggle at the top). Lets you set the copy SCALE %
(1-100) per tracker — written to scale_<name>.txt, which the bot reads and uses to
re-scale ALL trades every cycle (only counting those worth >= $1 at that scale). Also a
manual Refresh button. Otherwise read-only — it never touches the bots' Excel or keys.

Env:  DASHBOARD_DATA_DIR  where the state_*.json + scale_*.txt files live
      DASHBOARD_PORT      port (default 8080)
      DASHBOARD_USER/PASS optional HTTP basic-auth (set both to require login)
"""
from __future__ import annotations

import glob
import json
import os

from flask import Flask, Response, redirect, request

DATA_DIR = os.environ.get("DASHBOARD_DATA_DIR", os.path.dirname(os.path.abspath(__file__)))

# Site password — PASSWORD ONLY (the username is ignored, leave it blank). Hardcoded
# default so it survives redeploys / systemd-service resets; can still be overridden with
# the DASHBOARD_PASS env var. The old "CHANGE_ME_NOW" placeholder counts as unset.
_envpw = os.environ.get("DASHBOARD_PASS", "").strip()
SITE_PASSWORD = _envpw if _envpw and _envpw != "CHANGE_ME_NOW" else "123"
REALM = "Polymarket Copytrader - password only (leave username blank)"

app = Flask(__name__)


def _auth_ok() -> bool:
    a = request.authorization
    return bool(a and a.password == SITE_PASSWORD)  # username ignored


def _need_auth() -> Response:
    return Response("Password required", 401, {"WWW-Authenticate": f'Basic realm="{REALM}"'})


def _load_states() -> dict[str, dict]:
    out: dict[str, dict] = {}
    for p in sorted(glob.glob(os.path.join(DATA_DIR, "state_*.json"))):
        try:
            with open(p) as fh:
                s = json.load(fh)
            out[s.get("name", os.path.basename(p))] = s
        except (OSError, ValueError):
            continue
    return out


def _scale_path(name: str) -> str:
    return os.path.join(DATA_DIR, f"scale_{name}.txt")


def _read_scale(name: str) -> float | None:
    try:
        v = float(open(_scale_path(name)).read().strip())
        return max(0.01, min(100.0, v))
    except (OSError, ValueError):
        return None


def _write_scale(name: str, v: float) -> None:
    with open(_scale_path(name), "w") as fh:
        fh.write(f"{v:g}")


def _maxdelta_path(name: str) -> str:
    return os.path.join(DATA_DIR, f"maxdelta_{name}.txt")


def _read_maxdelta(name: str) -> float:
    """Max entry-delta % view filter (0 = no limit)."""
    try:
        v = float(open(_maxdelta_path(name)).read().strip())
        return v if v > 0 else 0.0
    except (OSError, ValueError):
        return 0.0


def _write_maxdelta(name: str, v: float) -> None:
    with open(_maxdelta_path(name), "w") as fh:
        fh.write(f"{v:g}")


def _live_path(name: str) -> str:
    return os.path.join(DATA_DIR, f"live_{name}.txt")


def _read_live(name: str) -> bool:
    try:
        return open(_live_path(name)).read().strip().lower() == "on"
    except OSError:
        return False


def _write_live(name: str, on: bool) -> None:
    with open(_live_path(name), "w") as fh:
        fh.write("on" if on else "off")


def _money(x) -> str:
    try:
        return f"${float(x):,.2f}"
    except (TypeError, ValueError):
        return str(x)


def _delta3(x) -> str:
    """Unsigned price difference to 3dp, e.g. 0.032 — delta is a magnitude (always >= 0)."""
    try:
        return f"{abs(float(x)):.3f}"
    except (TypeError, ValueError):
        return str(x)


def _pct1(x) -> str:
    """Percent to 1dp, e.g. 12.5% — for delta percentage."""
    try:
        return f"{abs(float(x)):.1f}%"
    except (TypeError, ValueError):
        return ""


def _cls(x) -> str:
    try:
        return "pos" if float(x) > 0 else ("neg" if float(x) < 0 else "")
    except (TypeError, ValueError):
        return ""


CSS = """
* { box-sizing: border-box; }
body { margin:0; background:#ffffff; color:#1f2328; font:14px/1.5 -apple-system,Segoe UI,Roboto,sans-serif; }
.wrap { max-width:1200px; margin:0 auto; padding:24px; }
h1 { font-size:18px; margin:0 0 2px; }
.sub { color:#656d76; font-size:12px; margin-bottom:14px; }
.live { color:#bc4c00; font-weight:700; }
.topbar { display:flex; gap:14px; align-items:center; flex-wrap:wrap; margin-bottom:6px; }
.toggle { display:inline-flex; border:1px solid #d0d7de; border-radius:9px; overflow:hidden; }
.toggle a { padding:7px 18px; text-decoration:none; color:#1f2328; background:#fff; font-weight:600; font-size:13px; }
.toggle a.active { background:#0969da; color:#fff; }
.toggle a .livedot { color:#cf222e; font-size:9px; font-weight:800; margin-left:5px; vertical-align:middle; }
.toggle a.active .livedot { color:#ffd7d5; }
.controls { display:flex; gap:10px; align-items:center; flex-wrap:wrap; margin:12px 0; }
.btn { padding:6px 12px; border:1px solid #d0d7de; border-radius:8px; background:#f6f8fa; cursor:pointer; font:inherit; font-size:13px; color:#1f2328; text-decoration:none; }
.btn:hover { background:#eaeef2; }
form.scale { display:inline-flex; gap:6px; align-items:center; margin:0; }
input[type=number] { width:64px; padding:5px 6px; border:1px solid #d0d7de; border-radius:6px; font:inherit; }
.hint { color:#8c959f; font-size:11px; }
.cards { display:grid; grid-template-columns:repeat(auto-fit, minmax(135px, 1fr)); gap:10px; margin-bottom:14px; }
.card.peak { background:#fff8c5; border-color:#d4a72c; }
.daily { display:flex; gap:6px; flex-wrap:wrap; align-items:center; margin-bottom:14px; }
.day { display:inline-block; background:#f6f8fa; border:1px solid #d0d7de; border-radius:6px; padding:4px 9px; text-align:center; text-decoration:none; color:inherit; cursor:pointer; }
.day:hover { border-color:#0969da; }
.day.active { border-color:#0969da; background:#ddf4ff; }
.day .d { color:#656d76; font-size:10px; }
.day .v { font-size:12px; font-weight:700; }
.banner { background:#ddf4ff; border:1px solid #0969da; border-radius:8px; padding:8px 12px; margin-bottom:12px; font-size:13px; }
.liveswitch { padding:7px 16px; border-radius:9px; border:1px solid #d0d7de; font:inherit; font-weight:700; font-size:13px; cursor:pointer; }
.liveswitch.off { background:#f6f8fa; color:#656d76; }
.liveswitch.on { background:#cf222e; color:#fff; border-color:#cf222e; }
.livewarn { background:#cf222e; color:#fff; padding:9px 13px; border-radius:8px; margin-bottom:12px; font-weight:700; }
.card { background:#f6f8fa; border:1px solid #d0d7de; border-radius:10px; padding:12px 14px; }
.card .label { color:#656d76; font-size:10px; letter-spacing:.04em; text-transform:uppercase; }
.card .val { font-size:20px; font-weight:700; margin-top:4px; }
.card.green { background:#dafbe1; border-color:#2da44e; }
.card.red { background:#ffebe9; border-color:#cf222e; }
.tablewrap { overflow-x:auto; -webkit-overflow-scrolling:touch; border-radius:10px; }
/* Full-bleed: break the table out of the centered .wrap so it spans the whole screen. */
.fullbleed { width:100vw; margin-left:calc(50% - 50vw); margin-right:calc(50% - 50vw); }
.fullbleed.tablewrap { border-radius:0; }
.fullbleed table { border-radius:0; border-left:none; border-right:none; }
.tabletools { width:100vw; margin-left:calc(50% - 50vw); display:flex; justify-content:flex-end; padding:0 12px; margin-bottom:6px; }
.dropdown { position:relative; }
.colmenu { display:none; position:absolute; right:0; top:34px; z-index:30; background:#fff; border:1px solid #d0d7de; border-radius:8px; box-shadow:0 6px 18px rgba(0,0,0,.12); padding:6px; min-width:172px; }
.colmenu.open { display:block; }
.colmenu .mh { font-size:10px; text-transform:uppercase; letter-spacing:.04em; color:#8c959f; padding:4px 8px 6px; }
.colmenu label { display:flex; gap:8px; align-items:center; padding:5px 8px; font-size:13px; cursor:pointer; white-space:nowrap; border-radius:6px; }
.colmenu label:hover { background:#f6f8fa; }
.colmenu input { margin:0; }
.colmenu .exitem { display:block; padding:7px 9px; font-size:13px; color:#1f2328; text-decoration:none; border-radius:6px; white-space:nowrap; }
.colmenu .exitem:hover { background:#f6f8fa; }
.colmenu .exitem .exsub { display:block; font-size:11px; color:#8c959f; }
table { width:100%; border-collapse:collapse; background:#ffffff; border:1px solid #d0d7de; border-radius:10px; overflow:hidden; }
th,td { padding:7px 10px; text-align:right; border-bottom:1px solid #d8dee4; white-space:nowrap; }
th { background:#f6f8fa; color:#656d76; font-size:11px; text-transform:uppercase; position:sticky; top:0; }
td.l,th.l { text-align:left; }
tr:hover td { background:#f6f8fa; }
.pos { color:#1a7f37; } .neg { color:#cf222e; }
.tag { font-size:10px; padding:1px 7px; border-radius:10px; }
.tag.OPEN { background:#dafbe1; color:#1a7f37; } .tag.RESOLVED { background:#eaeef2; color:#656d76; }
.foot { color:#8c959f; font-size:11px; margin-top:18px; }

/* --- Mobile (iPhone-width and below): reflow so it fits neatly --- */
@media (max-width: 480px) {
  .wrap { padding:12px; }
  .topbar { flex-direction:column; align-items:flex-start; gap:8px; }
  .topbar h1 { font-size:16px; }
  .toggle a { padding:8px 22px; }              /* big tap targets */
  .cards { grid-template-columns:repeat(2,1fr); gap:8px; }
  .card { padding:10px 11px; }
  .card .val { font-size:16px; }
  .controls { gap:8px; }
  .hint { display:none; }                       /* hide long hint to save space */
  .daily { gap:5px; }
  table { font-size:12px; }
  th, td { padding:6px 8px; }
  .banner { font-size:12px; }
}
"""


# Trade-list columns: (key, header label, left-aligned?). `key` is stamped as data-col on
# every <th>/<td> so the column-chooser JS can show/hide it client-side. ALL columns are
# always rendered (data is never dropped) — hiding is purely visual and remembered per
# browser via localStorage.
COLUMNS = [
    ("market", "Market", True),
    ("outcome", "Outcome", True),
    ("status", "Status", False),
    ("opened", "Opened", True),
    ("size", "Size", False),
    ("our_entry", "Our Entry", False),
    ("whale_entry", "Whale Entry", False),
    ("delta", "Delta", False),
    ("delta_pct", "Delta %", False),
    ("cur_price", "Cur Price", False),
    ("value", "Value", False),
    ("pnl", "P&L", False),
    ("lag", "Lag s", False),
]


def _row_cells(r: dict) -> dict[str, str]:
    st = r.get("status", "")
    return {
        "market": f'<td class="l" data-col="market">{r.get("market_title","")}</td>',
        "outcome": f'<td class="l" data-col="outcome">{r.get("outcome","")}</td>',
        "status": f'<td data-col="status"><span class="tag {st}">{st}</span></td>',
        "opened": f'<td class="l" data-col="opened">{r.get("trade_time","")}</td>',
        "size": f'<td data-col="size">{r.get("net_paper_size","")}</td>',
        "our_entry": f'<td data-col="our_entry">{r.get("avg_entry_price","")}</td>',
        "whale_entry": f'<td data-col="whale_entry">{r.get("whale_entry","")}</td>',
        "delta": f'<td data-col="delta">{_delta3(r.get("delta", 0)) if r.get("delta", "") != "" else ""}</td>',
        "delta_pct": f'<td data-col="delta_pct">{_pct1(r.get("delta_pct")) if r.get("delta_pct", "") != "" else ""}</td>',
        "cur_price": f'<td data-col="cur_price">{r.get("current_price","")}</td>',
        "value": f'<td data-col="value">{_money(r.get("current_value",0))}</td>',
        "pnl": f'<td data-col="pnl" class="{_cls(r.get("pnl",0))}">{_money(r.get("pnl",0))}</td>',
        "lag": f'<td data-col="lag">{r.get("avg_lag_s","")}</td>',
    }


def _render_tracker(s: dict, scale_val: float, date_filter: str = "", maxdelta_val: float = 0.0) -> str:
    name = s.get("name", "?")
    live_mode = bool(s.get("live"))
    # In live mode show ONLY real placed orders; paper data stays computed in the
    # background (s["summary"]/["positions"]) so switching back is instant.
    if live_mode:
        summ = s.get("live_summary") or {}
        live = '<span class="live">LIVE</span>'
    else:
        summ = s.get("summary", {})
        live = "paper (dry-run)"

    def cardcls(x):
        return "green" if _cls(x) == "pos" else ("red" if _cls(x) == "neg" else "")
    tp = summ.get("total_pnl", 0)
    cards = [
        ("Total P&L", _money(tp), cardcls(tp)),
        ("Today's P&L", _money(summ.get("today_pnl", 0)), cardcls(summ.get("today_pnl", 0))),
        ("Realized · resolved", _money(summ.get("realized_pnl", 0)), cardcls(summ.get("realized_pnl", 0))),
        ("Unrealized · open", _money(summ.get("unrealized_pnl", 0)), cardcls(summ.get("unrealized_pnl", 0))),
        ("Avg Delta · |our−whale entry|", _delta3(summ.get("avg_delta", 0)), ""),
        ("Total Spent", _money(summ.get("total_cost", 0)), ""),
        ("Portfolio Value · open", _money(summ.get("portfolio_value", 0)), ""),
        ("Peak Open · cash to fund", _money(summ.get("peak_open_value", 0)), "peak"),
        ("Open / Resolved", f"{summ.get('open_count',0)} / {summ.get('resolved_count',0)}", ""),
    ]
    card_html = "".join(
        f'<div class="card {c}"><div class="label">{l}</div><div class="val">{v}</div></div>'
        for l, v, c in cards)

    date_q = f"&date={date_filter}" if date_filter else ""
    export_btn = (
        '<div class="dropdown" style="margin-left:auto">'
        '<button class="btn" onclick="toggleMenu(event,\'exportmenu\')">&darr; Export &#9662;</button>'
        '<div class="colmenu" id="exportmenu">'
        '<div class="mh">Export as Excel</div>'
        f'<a class="exitem" href="/export?t={name}{date_q}&kind=positions">Positions <span class="exsub">combined per event</span></a>'
        f'<a class="exitem" href="/export?t={name}{date_q}&kind=trades">All individual trades <span class="exsub">every fill</span></a>'
        '</div></div>')
    daily = summ.get("daily", [])[:14]
    if daily:
        allcls = "" if date_filter else " active"
        chips = f'<a class="day{allcls}" href="/?t={name}"><div class="d">ALL</div><div class="v">&mdash;</div></a>'
        chips += "".join(
            f'<a class="day{" active" if d == date_filter else ""}" href="/?t={name}&date={d}">'
            f'<div class="d">{d[5:]}</div><div class="v {_cls(p)}">{_money(p)}</div></a>'
            for d, p in daily)
        daily_html = (f'<div class="daily"><span class="sub">Daily realized:</span> '
                      f'{chips}{export_btn}</div>')
    else:
        daily_html = f'<div class="daily">{export_btn}</div>'

    rows = (s.get("live_positions") or []) if live_mode else s.get("positions", [])
    if date_filter:
        rows = [r for r in rows if r.get("trade_date") == date_filter]
    banner = ""
    if live_mode and not (s.get("live_positions") or []):
        banner = ('<div class="banner">No live orders placed yet — this view shows only '
                  'trades that were <b>actually executed</b> on the exchange. New copied '
                  'trades will appear here once they fill. (Paper trading keeps running in '
                  'the background — switch to Paper to see it.)</div>')
    if date_filter:
        banner = (f'<div class="banner">Showing positions opened on <b>{date_filter}</b> '
                  f'({len(rows)}) · <a href="/?t={name}">show all</a></div>')
    body = []
    for r in rows[:200]:
        cells = _row_cells(r)
        body.append("<tr>" + "".join(cells[k] for k, _, _ in COLUMNS) + "</tr>")
    note = f" (showing first 200 of {len(rows)})" if len(rows) > 200 else ""

    header_cells = "".join(
        f'<th data-col="{k}"{" class=\"l\"" if left else ""}>{label}</th>'
        for k, label, left in COLUMNS)
    col_menu = "".join(
        f'<label><input type="checkbox" data-colkey="{k}" checked '
        f'onchange="onColToggle()">{label}</label>' for k, label, _ in COLUMNS)
    tabletools = (
        '<div class="tabletools"><div class="dropdown">'
        '<button class="btn" onclick="toggleMenu(event,\'colmenu\')">&#9776; Columns &#9662;</button>'
        f'<div class="colmenu" id="colmenu"><div class="mh">Show columns</div>{col_menu}</div>'
        '</div></div>')

    if live_mode:
        controls = """
    <div class="controls">
      <button class="btn" onclick="location.reload()">&#8635; Refresh</button>
      <span class="hint">Live view: only orders actually placed on the exchange. The scale
        slider is a paper-only setting and doesn't affect live orders (those are fixed at
        execution). Switch to Paper to adjust scale or see simulated trades.</span>
    </div>"""
        meta = (f"Updated {summ.get('last_updated','')} · <b>LIVE orders only</b> · "
                f"Return {summ.get('total_pnl_pct',0):.2f}% · "
                f"{summ.get('order_count',0)} order(s){note}")
    else:
        controls = f"""
    <div class="controls">
      <form class="scale" method="post" action="/scale">
        <input type="hidden" name="tracker" value="{name}">
        <label>Scale %: <input id="scaleinput" type="number" name="scale" min="0.01" max="100" step="0.01" value="{scale_val:g}"></label>
        <button class="btn">Apply</button>
      </form>
      <button class="btn" onclick="location.reload()">&#8635; Refresh</button>
      <span class="hint">Decimals allowed (e.g. 0.1 = 0.1%). Recomputes ALL trades (only those &ge; $1 at that scale); applies within ~1 min.</span>
    </div>
    <div class="controls">
      <form class="scale" method="post" action="/maxdelta">
        <input type="hidden" name="tracker" value="{name}">
        <label>Max delta %: <input type="number" name="maxdelta" min="0" step="0.1" value="{maxdelta_val:g}" placeholder="no limit"></label>
        <button class="btn">Apply</button>
      </form>
      <span class="hint">0 = no limit. Hides positions whose entry is &gt; this % from the whale's. All trades stay recorded — clear it to bring them back.</span>
    </div>"""
        _md = summ.get('max_delta_pct', 0) or 0
        _mdtxt = f" · max &Delta; {_md:g}% ({summ.get('delta_filtered_count',0)} filtered)" if _md else ""
        meta = (f"Updated {summ.get('last_updated','')} · Scale {summ.get('scale_pct',100):g}%"
                f" · Return {summ.get('total_pnl_pct',0):.2f}% · {summ.get('priced_count',0)} counted,"
                f" {summ.get('hidden_count',0)} hidden &lt;$1{_mdtxt}{note}")

    return f"""
      <h1>@{name} <span class="sub">· {live}</span></h1>
      <div class="sub">{meta}</div>
      {controls}
      <div class="cards">{card_html}</div>
      {daily_html}
      {banner}
      {tabletools}
      <div class="tablewrap fullbleed"><table>
        <tr>{header_cells}</tr>
        {''.join(body)}
      </table></div>"""


def _build_export(name: str, s: dict, rows: list[dict], live_mode: bool, date_filter: str) -> bytes:
    """Render an .xlsx in memory: a P&L summary header + the aggregated positions table
    (one row per market/outcome — already combined across trades). Returns the file bytes."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    view = (s.get("live_summary") or {}) if live_mode else s.get("summary", {})

    def f(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return 0.0

    # P&L summary computed FROM the (possibly date-filtered) rows, so it always matches
    # exactly what's exported below.
    open_rows = [r for r in rows if r.get("status") == "OPEN"]
    res_rows = [r for r in rows if r.get("status") == "RESOLVED"]
    tp = sum(f(r.get("pnl")) for r in rows)
    realized = sum(f(r.get("pnl")) for r in res_rows)
    unreal = sum(f(r.get("pnl")) for r in open_rows)
    cost = sum(f(r.get("cost_basis")) for r in rows)
    value = sum(f(r.get("current_value")) for r in open_rows)
    tc = sum(f(r.get("cost_basis")) for r in rows)   # dollar-weighted (by cost basis)
    avg_delta = (sum(f(r.get("cost_basis")) * f(r.get("delta")) for r in rows) / tc) if tc > 1e-9 else 0.0

    CCY, PRICE, SIZE, PCT = "$#,##0.00", "0.000", "#,##0.0000", '0.0"%"'
    NAVY, BLUE, GREEN, RED, LIGHT = "1F3864", "2F5496", "2E7D32", "C62828", "F2F6FC"
    cols = [("Market", "l", 42), ("Outcome", "l", 22), ("Status", "c", 10), ("Opened", "l", 17),
            ("Size", "c", 12), ("Our Entry", "c", 10), ("Whale Entry", "c", 12), ("Delta", "c", 9),
            ("Delta %", "c", 9), ("Cur Price", "c", 10), ("Value", "c", 12), ("P&L", "c", 12),
            ("Lag s", "c", 8)]
    ncols = len(cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "Positions"
    period = date_filter or "All days"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, f"Polymarket Copytrader  ·  @{name}  ·  {'LIVE' if live_mode else 'paper'}  ·  {period}")
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sub = ws.cell(2, 1, f"Updated {view.get('last_updated','')}   ·   {len(rows)} positions "
                        f"({len(open_rows)} open / {len(res_rows)} resolved)")
    sub.font = Font(size=10, italic=True, color="D9E1F2")
    for r in (1, 2):
        for c in range(1, ncols + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 24

    # Summary cards (label row 3, value row 4), distributed across the columns.
    cards = [("TOTAL P&L", tp, CCY), ("REALIZED", realized, CCY), ("UNREALIZED", unreal, CCY),
             ("AVG DELTA", avg_delta, PRICE), ("TOTAL SPENT", cost, CCY),
             ("PORTFOLIO VALUE", value, CCY), ("OPEN / RESOLVED", f"{len(open_rows)} / {len(res_rows)}", None)]
    base, rem = divmod(ncols, len(cards))
    spans, start = [], 1
    for i in range(len(cards)):
        w = base + (1 if i < rem else 0)
        spans.append((start, start + w - 1))
        start += w
    for (label, val, fmt), (c0, c1) in zip(cards, spans):
        ws.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c1)
        ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c1)
        lc = ws.cell(3, c0, label)
        lc.font = Font(bold=True, size=9, color="FFFFFF")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        fill = GREEN if (fmt == CCY and isinstance(val, (int, float)) and val > 0) else \
            (RED if (fmt == CCY and isinstance(val, (int, float)) and val < 0) else BLUE)
        vc = ws.cell(4, c0, val)
        vc.font = Font(bold=True, size=13, color="FFFFFF")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            vc.number_format = fmt
        for r in (3, 4):
            for c in range(c0, c1 + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[4].height = 24

    hr = 6
    for i, (label, _align, width) in enumerate(cols, start=1):
        cell = ws.cell(hr, i, label)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    fmts = [None, None, None, None, SIZE, PRICE, PRICE, PRICE, PCT, PRICE, CCY, CCY, None]
    for ri, r in enumerate(rows):
        row = hr + 1 + ri
        band = PatternFill("solid", fgColor=LIGHT) if ri % 2 else None
        vals = [r.get("market_title", ""), r.get("outcome", ""), r.get("status", ""),
                r.get("trade_time", ""), f(r.get("net_paper_size")), f(r.get("avg_entry_price")),
                f(r.get("whale_entry")), f(r.get("delta")), f(r.get("delta_pct")),
                f(r.get("current_price")), f(r.get("current_value")), f(r.get("pnl")),
                r.get("avg_lag_s", "")]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row, ci, v)
            if fmt:
                cell.number_format = fmt
            if band:
                cell.fill = band
            cell.alignment = Alignment(horizontal="left" if cols[ci - 1][1] == "l" else "center")
    ws.freeze_panes = f"A{hr + 1}"
    ws.sheet_view.showGridLines = False

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _excel_path_for(s: dict, name: str) -> str:
    """Locate a tracker's Excel trade-log in DATA_DIR. Prefer the filename the bot records
    in state (excel_file); fall back to common names / any matching .xlsx, so export works
    even before the bot has written a state with that field."""
    candidates = []
    if s.get("excel_file"):
        candidates.append(s["excel_file"])
    candidates += [f"{name}_paper_trades.xlsx", f"{name.lower()}_paper_trades.xlsx"]
    for fn in candidates:
        p = os.path.join(DATA_DIR, fn)
        if fn.lower().endswith(".xlsx") and os.path.exists(p):
            return p
    # Last resort: any .xlsx in the data dir whose name mentions this tracker.
    for p in glob.glob(os.path.join(DATA_DIR, "*.xlsx")):
        if name.lower() in os.path.basename(p).lower():
            return p
    return ""


def _build_trades_export(name: str, path: str, date_filter: str, live_mode: bool, s: dict) -> bytes:
    """Render an .xlsx of EVERY individual trade (one row per fill) from the bot's Trades
    sheet, read-only. Crucially, P&L / cost / current price are RECOMPUTED live from the
    current scale and the current prices in the positions state — NOT read from the Trades
    sheet's stale mark columns — so the totals reconcile with the Positions export. Cost &
    P&L are shown only for 'counted' trades (priced and >= $1 at the current scale), exactly
    the set the positions view aggregates, so the per-row values sum to the totals."""
    from datetime import datetime, timezone
    from io import BytesIO

    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not path or not os.path.exists(path):
        raise FileNotFoundError("trade log not found")

    src = load_workbook(path, read_only=True, data_only=True)
    try:
        ws_in = src["Trades"]
        it = ws_in.iter_rows(values_only=True)
        header = list(next(it))
        idx = {h: i for i, h in enumerate(header) if h}
        raw = [vals for vals in it if vals and vals[0] is not None]
    finally:
        src.close()

    def g(vals, key):
        i = idx.get(key)
        return vals[i] if (i is not None and i < len(vals)) else ""

    def numf(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return 0.0

    def tfmt(ts, fmt):
        try:
            return datetime.fromtimestamp(int(float(ts)), tz=timezone.utc).strftime(fmt)
        except (TypeError, ValueError, OSError, OverflowError):
            return ""

    # Current prices + scale from the SAME state the Positions view uses, so totals match.
    view = (s.get("live_summary") or {}) if live_mode else (s.get("summary") or {})
    pos = (s.get("live_positions") or []) if live_mode else (s.get("positions") or [])
    price_map: dict[str, float] = {}
    for p in pos:
        tok = str(p.get("token_id", ""))
        cp = p.get("current_price")
        if tok and cp not in (None, ""):
            try:
                price_map[tok] = float(cp)
            except (TypeError, ValueError):
                pass
    try:
        scale_frac = float(view.get("scale_pct") or 100) / 100.0
    except (TypeError, ValueError):
        scale_frac = 1.0
    MIN_USD = 1.0

    def calc(v):
        """Returns (side, size, entry, cost, cur, pnl, counted) recomputed live."""
        token = str(g(v, "token_id") or "")
        side = str(g(v, "side") or "BUY").upper()
        cur = price_map.get(token)            # None if this token isn't a priced position
        if live_mode:
            entry = numf(g(v, "live_avg_price")) or numf(g(v, "paper_entry_price"))
            filled = numf(g(v, "live_filled"))
            size = filled if filled > 0 else numf(g(v, "rn1_size")) * numf(g(v, "scale_ratio"))
            counted = cur is not None and size > 0 and entry > 0
        else:
            entry = numf(g(v, "paper_entry_price"))
            size = numf(g(v, "rn1_size")) * scale_frac
            counted = cur is not None and (size * entry) >= MIN_USD
        cost = size * entry
        pnl = None if cur is None else (size * (cur - entry) if side == "BUY" else size * (entry - cur))
        return side, size, entry, cost, cur, pnl, counted

    sel, calcs = [], []
    counted_n = 0
    total_cost = total_pnl = 0.0
    for v in raw:
        if date_filter and tfmt(g(v, "trade_ts"), "%Y-%m-%d") != date_filter:
            continue
        # Only trades we caught LIVE i.e. within the copy window (matches the positions view;
        # missed/backfilled trades with a huge lag are forgotten).
        try:
            _dlag = float(g(v, "detect_lag_s"))
        except (TypeError, ValueError):
            _dlag = 1e9
        if _dlag > 120:
            continue
        if live_mode:
            ls, oid = str(g(v, "live_status") or ""), str(g(v, "live_order_id") or "")
            if not (ls.startswith("LIVE") and oid.strip()):
                continue
        c = calc(v)
        sel.append(v)
        calcs.append(c)
        side, size, entry, cost, cur, pnl, counted = c
        if counted:
            counted_n += 1
            if side == "BUY":
                total_cost += cost
            if pnl is not None:
                total_pnl += pnl

    CCY, PRICE, SIZE, PCT = "$#,##0.00", "0.000", "#,##0.0000", '0.0"%"'
    NAVY, BLUE, GREEN, RED, LIGHT = "1F3864", "2F5496", "2E7D32", "C62828", "F2F6FC"
    cols = [("Trade Time", "l", 17, None), ("Market", "l", 42, None), ("Outcome", "l", 20, None),
            ("Side", "c", 7, None), ("Size", "c", 12, SIZE), ("Our Entry", "c", 10, PRICE),
            ("Whale Price", "c", 11, PRICE), ("Delta %", "c", 9, PCT), ("Cost", "c", 12, CCY),
            ("Cur Price", "c", 10, PRICE), ("P&L", "c", 12, CCY), ("Lag s", "c", 8, None),
            ("Source", "c", 9, None), ("Status", "l", 16, None)]
    ncols = len(cols)

    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    period = date_filter or "All days"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, f"Polymarket Copytrader  ·  @{name}  ·  {'LIVE orders' if live_mode else 'paper'}  ·  "
                      f"individual trades  ·  {period}")
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sub = ws.cell(2, 1, f"Updated {view.get('last_updated','')}   ·   {len(sel)} trades   ·   "
                        f"{counted_n} counted (priced & ≥ $1 at current scale)   ·   "
                        f"P&L/cost shown for counted trades only, to match the Positions view")
    sub.font = Font(size=10, italic=True, color="D9E1F2")
    for r in (1, 2):
        for c in range(1, ncols + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 24

    cards = [("TRADES", f"{counted_n} / {len(sel)}", None), ("TOTAL COST", total_cost, CCY),
             ("TOTAL P&L", total_pnl, CCY)]
    base, rem = divmod(ncols, len(cards))
    spans, start = [], 1
    for i in range(len(cards)):
        w = base + (1 if i < rem else 0)
        spans.append((start, start + w - 1))
        start += w
    for (label, val, fmt), (c0, c1) in zip(cards, spans):
        ws.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c1)
        ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c1)
        lc = ws.cell(3, c0, label)
        lc.font = Font(bold=True, size=9, color="FFFFFF")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        fill = GREEN if (fmt == CCY and isinstance(val, (int, float)) and val > 0) else \
            (RED if (fmt == CCY and isinstance(val, (int, float)) and val < 0) else BLUE)
        vc = ws.cell(4, c0, val)
        vc.font = Font(bold=True, size=13, color="FFFFFF")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            vc.number_format = fmt
        for r in (3, 4):
            for c in range(c0, c1 + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[4].height = 22

    hr = 6
    for i, (label, _a, width, _fmt) in enumerate(cols, start=1):
        cell = ws.cell(hr, i, label)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    for ri, (v, c) in enumerate(zip(sel, calcs)):
        side, size, entry, cost, cur, pnl, counted = c
        whale_p = numf(g(v, "rn1_price"))
        dpct = round(abs(entry - whale_p) / whale_p * 100, 2) if whale_p > 1e-9 else 0.0
        vals = [tfmt(g(v, "trade_ts"), "%Y-%m-%d %H:%M"), g(v, "market_title"), g(v, "outcome"),
                side, round(size, 6), round(entry, 6), whale_p, dpct,
                (round(cost, 4) if counted else ""), (round(cur, 6) if cur is not None else ""),
                (round(pnl, 4) if (counted and pnl is not None) else ""),
                g(v, "detect_lag_s"), g(v, "source"), g(v, "live_status")]
        row = hr + 1 + ri
        band = PatternFill("solid", fgColor=LIGHT) if ri % 2 else None
        for ci, ((_label, align, _w, fmt), val) in enumerate(zip(cols, vals), start=1):
            cell = ws.cell(row, ci, val)
            if fmt and isinstance(val, (int, float)):
                cell.number_format = fmt
            if band:
                cell.fill = band
            cell.alignment = Alignment(horizontal="left" if align == "l" else "center")
    ws.freeze_panes = f"A{hr + 1}"
    ws.sheet_view.showGridLines = False

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.route("/export")
def export():
    if not _auth_ok():
        return _need_auth()
    states = _load_states()
    name = request.args.get("t", "")
    if name not in states:
        return Response("Unknown tracker", 404)
    s = states[name]
    live_mode = bool(s.get("live"))
    date_filter = request.args.get("date", "")
    kind = request.args.get("kind", "positions")

    if kind == "trades":
        path = _excel_path_for(s, name)
        if not path:
            return Response("Trade log file not found for this tracker yet — try again in "
                            "a minute (the bot writes it each cycle).", 404)
        try:
            data = _build_trades_export(name, path, date_filter, live_mode, s)
        except Exception as exc:
            return Response(f"Could not read trade log: {exc}", 500)
        fname = f"{name}_trades_{date_filter or 'all'}.xlsx"
    else:
        rows = (s.get("live_positions") or []) if live_mode else s.get("positions", [])
        if date_filter:
            rows = [r for r in rows if r.get("trade_date") == date_filter]
        data = _build_export(name, s, rows, live_mode, date_filter)
        fname = f"{name}_positions_{date_filter or 'all'}.xlsx"

    return Response(data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.route("/live", methods=["POST"])
def set_live():
    if not _auth_ok():
        return _need_auth()
    name = request.form.get("tracker", "").strip()
    if name:
        _write_live(name, request.form.get("on", "0") == "1")
    return redirect(f"/?t={name}")


@app.route("/scale", methods=["POST"])
def set_scale():
    if not _auth_ok():
        return _need_auth()
    name = request.form.get("tracker", "").strip()
    try:
        v = max(0.01, min(100.0, float(request.form.get("scale", "100"))))
    except ValueError:
        v = 100.0
    if name:
        _write_scale(name, v)
    return redirect(f"/?t={name}")


@app.route("/maxdelta", methods=["POST"])
def set_maxdelta():
    if not _auth_ok():
        return _need_auth()
    name = request.form.get("tracker", "").strip()
    try:
        v = max(0.0, float(request.form.get("maxdelta", "0")))
    except ValueError:
        v = 0.0
    if name:
        _write_maxdelta(name, v)
    return redirect(f"/?t={name}")


@app.route("/")
def index():
    if not _auth_ok():
        return _need_auth()
    states = _load_states()
    if not states:
        inner, toggle, livebtn, livewarn = "<p>No data yet — waiting for the bot's first cycle.</p>", "", "", ""
    else:
        names = list(states)
        active = request.args.get("t") or names[0]
        if active not in states:
            active = names[0]
        tabs = []
        for n in names:
            dot = ' <span class="livedot">● LIVE</span>' if states[n].get("live") else ""
            cls = "active" if n == active else ""
            tabs.append(f'<a href="/?t={n}" class="{cls}">@{n}{dot}</a>')
        toggle = '<div class="toggle">' + "".join(tabs) + "</div>"
        s = states[active]
        scale_val = _read_scale(active)
        if scale_val is None:
            scale_val = float(s.get("summary", {}).get("scale_pct", 100) or 100)
        date_filter = request.args.get("date", "")
        maxdelta_val = _read_maxdelta(active)
        inner = _render_tracker(s, scale_val, date_filter, maxdelta_val)
        # Live/paper toggle (top-right). Turning it ON asks for confirmation.
        live = bool(s.get("live"))
        livebtn = (f'<form method="post" action="/live" onsubmit="return confirmLive(this)" '
                   f'style="margin-left:auto">'
                   f'<input type="hidden" name="tracker" value="{active}">'
                   f'<input type="hidden" name="on" value="{0 if live else 1}">'
                   f'<button class="liveswitch {"on" if live else "off"}">'
                   f'{"● LIVE" if live else "○ PAPER"}</button></form>')
        livewarn = (f'<div class="livewarn">⚠️ LIVE TRADING ACTIVE for @{active} — real '
                    f'orders are being placed for new copied trades.</div>') if live else ""

    return f"""<!doctype html><html><head><meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Polymarket Copytrader</title><style>{CSS}</style>
    <script>
      // Auto-refresh every 30s, but never while you're typing in the scale box or have
      // the column menu open (so a refresh can't yank the menu out from under you).
      setInterval(function() {{
        var el = document.getElementById('scaleinput');
        if (document.querySelector('.colmenu.open')) return;   // any dropdown open
        if (el && document.activeElement === el) return;
        location.reload();
      }}, 30000);
      function confirmLive(f) {{
        if (f.on.value === '1') return confirm(
          '⚠️ Switch to LIVE TRADING?\\n\\nReal orders with REAL MONEY will be placed for new copied trades. Only do this if you have funded the account and configured a key.');
        return true;
      }}
      // --- Column chooser. Hidden columns persist per-browser in localStorage, so the
      // choice survives the 30s auto-refresh. Data is always rendered; we only hide it. ---
      function colHidden() {{
        try {{ return JSON.parse(localStorage.getItem('pmHiddenCols') || '[]'); }}
        catch (e) {{ return []; }}
      }}
      function applyCols() {{
        var hidden = colHidden();
        document.querySelectorAll('[data-col]').forEach(function(el) {{
          el.style.display = hidden.indexOf(el.getAttribute('data-col')) >= 0 ? 'none' : '';
        }});
      }}
      function onColToggle() {{
        var hidden = [];
        document.querySelectorAll('#colmenu input[data-colkey]').forEach(function(cb) {{
          if (!cb.checked) hidden.push(cb.getAttribute('data-colkey'));
        }});
        localStorage.setItem('pmHiddenCols', JSON.stringify(hidden));
        applyCols();
      }}
      function toggleMenu(e, id) {{
        e.stopPropagation();
        var target = document.getElementById(id);
        document.querySelectorAll('.colmenu').forEach(function(m) {{
          if (m !== target) m.classList.remove('open');   // only one open at a time
        }});
        if (target) target.classList.toggle('open');
      }}
      document.addEventListener('click', function(e) {{
        document.querySelectorAll('.colmenu.open').forEach(function(m) {{
          if (!m.contains(e.target)) m.classList.remove('open');
        }});
      }});
      document.addEventListener('DOMContentLoaded', function() {{
        var hidden = colHidden();
        document.querySelectorAll('#colmenu input[data-colkey]').forEach(function(cb) {{
          cb.checked = hidden.indexOf(cb.getAttribute('data-colkey')) < 0;
        }});
        applyCols();
      }});
    </script></head>
    <body><div class="wrap">
      <div class="topbar">
        <h1 style="margin:0">Polymarket Copytrader</h1>
        {toggle}
        {livebtn}
      </div>
      <div class="sub">switch trader above · toggle live/paper top-right · auto-refreshes every 30s</div>
      {livewarn}
      {inner}
      <div class="foot">Paper trading unless a tracker shows LIVE. Not financial advice.</div>
    </div></body></html>"""


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("DASHBOARD_PORT", "8080")))
