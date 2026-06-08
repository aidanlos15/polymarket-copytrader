"""Local Excel (.xlsx) datastore for the paper-trade bot, with a styled dashboard.

A single workbook on disk is the source of truth — the bot keeps no other state, so
it's restart-safe (already-logged trades are de-duplicated by reading the file back).

  Positions  (first tab) a pinned SUMMARY dashboard at the top + per-token holdings,
             rewritten and re-priced every cycle. Frozen panes keep the summary + table
             header visible while you scroll.
  Trades     append-only log of every copied paper trade, marked to market each cycle.

The workbook is held in memory and saved after each mutation. If the file is open in
Excel, saving may fail with PermissionError — close it (or saving will retry next cycle).
"""
from __future__ import annotations

import os
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config

# NOTE: keep this physical order stable — existing rows on disk are stored in it.
TRADES_HEADER = [
    "trade_id", "detected_at", "trade_ts", "market_title", "slug", "outcome",
    "side", "token_id", "condition_id", "rn1_size", "rn1_price", "scale_ratio",
    "paper_size", "paper_entry_price", "paper_cost", "tx_hash",
    "current_price", "current_value", "pnl", "pnl_updated", "detect_lag_s",
    "live_status", "live_order_id", "live_filled", "live_avg_price",
]

# Aggregated per-token view. Rebuilt from scratch each cycle, so order is free to be
# reader-friendly: the human-readable columns first, long ID hashes last.
POSITIONS_HEADER = [
    "market_title", "outcome", "status", "net_paper_size", "total_bought",
    "avg_entry_price", "cost_basis", "current_price", "current_value", "pnl", "pnl_pct",
    "price_source", "last_updated", "token_id", "condition_id",
]

# Per-cycle re-written columns on the Trades sheet (paper_size/cost change with scale).
TRADE_MARK_COLS = ["paper_size", "paper_cost", "current_price", "current_value",
                   "pnl", "pnl_updated"]

# The editable "SCALE %" input cell on the Positions sheet (value the user types, 1-100).
SCALE_LABEL_CELL = "A5"
SCALE_INPUT_CELL = "B5"

# --- styling palette --------------------------------------------------------
NAVY = "1F3864"
BLUE = "2F5496"
LIGHT = "F2F6FC"
GREEN = "2E7D32"
RED = "C62828"
GREEN_TXT = "1B7F1B"
RED_TXT = "C62828"

_CCY = '$#,##0.00'        # money
_PRICE = '0.000'         # 0..1 probabilities
_SIZE = '#,##0.0000'     # share sizes
_PCT = '0.00"%"'         # already-scaled percent

# header name -> number format
_NUMFMT = {
    "cost_basis": _CCY, "current_value": _CCY, "pnl": _CCY, "paper_cost": _CCY,
    "avg_entry_price": _PRICE, "current_price": _PRICE, "paper_entry_price": _PRICE,
    "rn1_price": _PRICE, "pnl_pct": _PCT,
    "net_paper_size": _SIZE, "total_bought": _SIZE, "paper_size": _SIZE,
    "rn1_size": _SIZE, "scale_ratio": _SIZE, "detect_lag_s": '#,##0',
    "live_avg_price": _PRICE, "live_filled": _SIZE,
}

# Trades columns to hide (long hashes / redundant) and friendly widths.
_TRADES_HIDE = {"trade_id", "slug", "token_id", "condition_id", "tx_hash", "trade_ts",
                "live_order_id"}
_WIDTHS = {
    "market_title": 42, "outcome": 20, "side": 7, "detected_at": 19, "pnl_updated": 19,
    "last_updated": 19, "price_source": 13, "token_id": 22, "condition_id": 22,
    "detect_lag_s": 12, "status": 11,
}
_DEFAULT_WIDTH = 13

_THIN = Side(style="thin", color="D6DCE4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


class ExcelClient:
    def __init__(self) -> None:
        self.path = config.EXCEL_PATH
        if os.path.exists(self.path):
            self._wb = load_workbook(self.path)
        else:
            self._wb = Workbook()
            self._wb.remove(self._wb.active)
        self._trades = self._ensure_ws("Trades", TRADES_HEADER, enforce_header=True)
        self._positions = self._ensure_ws("Positions", POSITIONS_HEADER, enforce_header=False)
        # Positions first, so the file opens on the dashboard.
        self._wb._sheets.remove(self._positions)
        self._wb._sheets.insert(0, self._positions)
        self._wb.active = 0
        self._save()

    def _ensure_ws(self, title: str, header: list[str], enforce_header: bool):
        if title in self._wb.sheetnames:
            ws = self._wb[title]
            if enforce_header and [c.value for c in ws[1]] != header:
                ws.delete_rows(1)
                ws.insert_rows(1)
                for i, h in enumerate(header, start=1):
                    ws.cell(row=1, column=i, value=h)
        else:
            ws = self._wb.create_sheet(title)
            if enforce_header:
                ws.append(header)
        return ws

    def _save(self) -> None:
        try:
            self._wb.save(self.path)
        except PermissionError:
            print(f"  [excel] cannot save — is '{self.path}' open in Excel? Will retry next cycle.")

    # --- Trades log ---------------------------------------------------------

    def read_processed_trade_ids(self) -> set[str]:
        ids: set[str] = set()
        for row in self._trades.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                ids.add(str(row[0]))
        return ids

    def read_scale_pct(self) -> float | None:
        """Read the user-editable SCALE % from the file ON DISK (so manual edits in
        Excel are picked up). Returns a float in [1, 100], or None if unset/invalid/
        unreadable (e.g. the file is locked open in Excel)."""
        if not os.path.exists(self.path):
            return None
        try:
            wb = load_workbook(self.path, read_only=True, data_only=True)
            try:
                if "Positions" not in wb.sheetnames:
                    return None
                val = wb["Positions"][SCALE_INPUT_CELL].value
            finally:
                wb.close()
            pct = float(val)
            return pct if 1.0 <= pct <= 100.0 else None
        except (OSError, ValueError, TypeError):
            return None

    def append_trade(self, row: dict[str, Any]) -> None:
        self._trades.append([row.get(k, "") for k in TRADES_HEADER])
        self._save()

    def read_all_trades(self) -> list[dict]:
        out: list[dict] = []
        for vals in self._trades.iter_rows(min_row=2, values_only=True):
            if vals[0] is None:
                continue
            out.append({h: vals[i] if i < len(vals) else "" for i, h in enumerate(TRADES_HEADER)})
        return out

    def update_trade_marks(self, marks_by_id: dict[str, dict]) -> None:
        """Write the per-cycle columns onto each Trades row (matched by trade_id) and
        hide rows that fall below the $1 minimum at the current scale, then re-style."""
        cols = {name: TRADES_HEADER.index(name) + 1 for name in TRADE_MARK_COLS}
        for row in range(2, self._trades.max_row + 1):
            tid = self._trades.cell(row=row, column=1).value
            if tid is None:
                continue
            mk = marks_by_id.get(str(tid))
            if not mk:
                continue
            for name, col in cols.items():
                self._trades.cell(row=row, column=col, value=mk.get(name, ""))
            # Hide the row when our scaled order is below the $1 minimum.
            self._trades.row_dimensions[row].hidden = bool(mk.get("_hidden"))
        self._style_trades()
        self._save()

    # --- Positions / dashboard ---------------------------------------------

    def write_positions(self, rows: list[dict[str, Any]], summary: dict[str, Any],
                        scale_pct: float = 1.0) -> None:
        """Rebuild the Positions sheet: a pinned summary dashboard on top, an editable
        SCALE % input, then the per-token table. Frozen panes keep it all visible."""
        ws = self._positions
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))
        if ws.max_row >= 1:
            ws.delete_rows(1, ws.max_row)
        ncols = len(POSITIONS_HEADER)

        # Row 1: title, Row 2: subtitle
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        t = ws.cell(1, 1, f"Polymarket Paper Copytrading  ·  tracking @{config.TARGET_NAME}")
        t.font = Font(bold=True, size=16, color="FFFFFF")
        t.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        sub = ws.cell(2, 1, f"Updated {summary.get('last_updated','')}   ·   "
                            f"Return {summary.get('total_pnl_pct',0):.2f}%   ·   "
                            f"{summary.get('priced_count',0)} trades priced, "
                            f"{summary.get('unpriced_count',0)} unpriced")
        sub.font = Font(size=10, italic=True, color="D9E1F2")
        sub.alignment = Alignment(horizontal="left", vertical="center")
        for r in (1, 2):
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=NAVY)
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 16

        # Rows 3-4: KPI cards (label on row 3, value on row 4)
        def signfill(v):
            return GREEN if v > 0 else (RED if v < 0 else BLUE)
        pnl = summary.get("total_pnl", 0) or 0
        today = summary.get("today_pnl", 0) or 0
        real = summary.get("realized_pnl", 0) or 0
        unreal = summary.get("unrealized_pnl", 0) or 0
        oc, rc = summary.get("open_count", 0), summary.get("resolved_count", 0)
        cards = [
            ("TOTAL P&L", pnl, _CCY, signfill(pnl)),
            ("TODAY'S P&L", today, _CCY, signfill(today)),
            ("REALIZED · resolved", real, _CCY, signfill(real)),
            ("UNREALIZED · open", unreal, _CCY, signfill(unreal)),
            ("TOTAL SPENT", summary.get("total_cost", 0), _CCY, BLUE),
            ("PORTFOLIO VALUE · open", summary.get("portfolio_value", 0), _CCY, BLUE),
            ("OPEN / RESOLVED", f"{oc} / {rc}", None, BLUE),
        ]
        # Distribute the columns across the cards (extra width to the leftmost cards).
        n = len(cards)
        base, rem = divmod(ncols, n)
        spans, start = [], 1
        for i in range(n):
            width = base + (1 if i < rem else 0)
            spans.append((start, start + width - 1))
            start += width
        for (label, value, fmt, fill), (c0, c1) in zip(cards, spans):
            ws.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c1)
            ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c1)
            lc = ws.cell(3, c0, label)
            lc.font = Font(bold=True, size=9, color="FFFFFF")
            lc.alignment = Alignment(horizontal="center", vertical="center")
            vc = ws.cell(4, c0, value)
            vc.font = Font(bold=True, size=15, color="FFFFFF")
            vc.alignment = Alignment(horizontal="center", vertical="center")
            if fmt:
                vc.number_format = fmt
            for r in (3, 4):
                for c in range(c0, c1 + 1):
                    ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 26

        # Row 5: editable SCALE % control. The bot reads B5 each cycle and resizes every
        # trade to it; trades whose scaled order is < $1 are hidden until you raise this.
        lab = ws.cell(5, 1, "SCALE %  →")
        lab.font = Font(bold=True, size=11, color="FFD966")
        lab.alignment = Alignment(horizontal="right", vertical="center")
        lab.fill = PatternFill("solid", fgColor="3A3000")
        inp = ws.cell(5, 2, round(scale_pct))
        inp.font = Font(bold=True, size=13, color="000000")
        inp.alignment = Alignment(horizontal="center", vertical="center")
        inp.fill = PatternFill("solid", fgColor="FFD966")   # yellow = editable
        inp.border = _BORDER
        inp.number_format = "0"
        note = ws.cell(5, 3, "← type 1–100, then save the file (applies on the next refresh)")
        note.font = Font(size=10, italic=True, color="8b949e")
        note.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=ncols)
        ws.row_dimensions[5].height = 22

        # Rows 6-7: per-day REALIZED P&L strip, most-recent-day first (date row 6, P&L row 7).
        daily = summary.get("daily", [])[:ncols - 1]      # already sorted newest-first
        dl = ws.cell(6, 1, "DAILY REALIZED →")
        dl.font = Font(bold=True, size=10, color="9CC3FF")
        dl.alignment = Alignment(horizontal="right", vertical="center")
        for j, (d, p) in enumerate(daily):
            col = 2 + j
            dc = ws.cell(6, col, d[5:])                   # MM-DD
            dc.font = Font(size=9, color="9CA3AF")
            dc.alignment = Alignment(horizontal="center")
            vc = ws.cell(7, col, round(p, 2))
            vc.number_format = _CCY
            vc.font = Font(bold=True, size=10,
                           color=GREEN_TXT if p > 0 else (RED_TXT if p < 0 else "9CA3AF"))
            vc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[6].height = 14
        ws.row_dimensions[7].height = 16

        # Row 8 spacer, Row 9 table header, Row 10+ data
        header_row = 9
        for i, h in enumerate(POSITIONS_HEADER, start=1):
            cell = ws.cell(header_row, i, h)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER
        ws.row_dimensions[header_row].height = 20

        for ri, r in enumerate(rows):
            excel_row = header_row + 1 + ri
            band = PatternFill("solid", fgColor=LIGHT) if ri % 2 else None
            for ci, h in enumerate(POSITIONS_HEADER, start=1):
                cell = ws.cell(excel_row, ci, r.get(h, ""))
                cell.border = _BORDER
                if band:
                    cell.fill = band
                if h in _NUMFMT:
                    cell.number_format = _NUMFMT[h]
                if h in ("market_title", "outcome", "price_source", "last_updated"):
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="center")

        self._apply_widths(ws, POSITIONS_HEADER)
        self._color_pnl(ws, POSITIONS_HEADER, ["pnl", "pnl_pct"], header_row + 1)
        ws.freeze_panes = f"A{header_row + 1}"  # pin summary + table header
        ws.sheet_view.showGridLines = False
        self._save()

    # --- styling helpers ----------------------------------------------------

    def _style_trades(self) -> None:
        ws = self._trades
        for i, h in enumerate(TRADES_HEADER, start=1):
            cell = ws.cell(1, i)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        for row in range(2, ws.max_row + 1):
            band = PatternFill("solid", fgColor=LIGHT) if row % 2 == 0 else None
            for ci, h in enumerate(TRADES_HEADER, start=1):
                cell = ws.cell(row, ci)
                if h in _NUMFMT:
                    cell.number_format = _NUMFMT[h]
                if band and cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = band

        self._apply_widths(ws, TRADES_HEADER, hide=_TRADES_HIDE)
        self._color_pnl(ws, TRADES_HEADER, ["pnl"], 2)
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

    def _apply_widths(self, ws, header, hide: set[str] = frozenset()) -> None:
        for i, h in enumerate(header, start=1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = _WIDTHS.get(h, _DEFAULT_WIDTH)
            if h in hide:
                ws.column_dimensions[letter].hidden = True

    def _color_pnl(self, ws, header, pnl_cols, first_data_row) -> None:
        """Green/red font for positive/negative P&L cells. Reset first to avoid
        accumulating duplicate rules across cycles."""
        ws.conditional_formatting = ConditionalFormattingList()
        last = ws.max_row
        if last < first_data_row:
            return
        for h in pnl_cols:
            if h not in header:
                continue
            letter = get_column_letter(header.index(h) + 1)
            rng = f"{letter}{first_data_row}:{letter}{last}"
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="greaterThan", formula=["0"], font=Font(color=GREEN_TXT, bold=True)))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="lessThan", formula=["0"], font=Font(color=RED_TXT, bold=True)))
